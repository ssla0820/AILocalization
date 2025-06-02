import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir)))

from bs4 import BeautifulSoup
from collections import OrderedDict
from chat.openai_api_chat import OpenaiAPIChat
from database.search_similar_pair import main as search_similar_pair_main
from pages.general_functions import get_relevant_specific_names, as_json_obj, InlineGroup, get_text_group_inline, load_specific_names, detect_file_encoding
from prompts.translate_prompts import *
from prompts.restruct_prompts import *
from translate.restruct import *
from review.review import *
import asyncio
import re
import math
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from config import translate_config as conf

def debug_process(
        source_text_index: str,
        source_text: str,
        relevant_specific_names: dict[str, str],
        relevant_pair_database: list[dict],
        prompt: str,
        response: str,
        output: dict[str, str]
):
    """
    Debugging function to print the translation process details.
    :param source_text_index: Index of the source text
    :param source_text: The original text to be translated
    :param relevant_specific_names: Specific names relevant for translation
    :param relevant_pair_database: Database of similar pairs for reference
    :param prompt: The prompt used for translation
    :param response: The response received from the translation API
    """
    # For Debugging: append source text/ relevant specific names/ relevant pair database/ prompt/ response to Excel file
    debug_file = 'debug.xlsx'
    try:
        # Check if file exists to determine if we need headers
        file_exists = os.path.isfile(debug_file)
        
        # Load workbook if it exists, otherwise create a new one
        if file_exists:
            wb = openpyxl.load_workbook(debug_file)
            ws = wb.active
            # Get next empty row
            next_row = ws.max_row + 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            # Add headers if creating a new file
            headers = ["Source Index", "Source Text", "Specific Names", "Similar Pairs", "Prompt", "Response", "Output"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            next_row = 2  # Start data from row 2
        
        # Add data to the next row
        ws.cell(row=next_row, column=1, value=str(source_text_index))
        ws.cell(row=next_row, column=2, value=str(source_text))
        ws.cell(row=next_row, column=3, value=str(relevant_specific_names))
        ws.cell(row=next_row, column=4, value=str(relevant_pair_database))
        ws.cell(row=next_row, column=5, value=str(prompt))
        ws.cell(row=next_row, column=6, value=str(response))
        ws.cell(row=next_row, column=7, value=str(output))
        
        # Apply word wrap for better readability in Excel
        for col in range(1, 8):
            cell = ws.cell(row=next_row, column=col)
            cell.alignment = Alignment(wrap_text=True)
        
        # Save the workbook
        wb.save(debug_file)
        
    except Exception as e:
        print(f"Warning: Could not save debug info to Excel: {e}")


def segment_groups_map(
        groups_map: dict[str, InlineGroup],
        max_token: int,
        token_counter: callable
) -> list[OrderedDict]:
    """
    Segments a map of inline groups based on token count, hoping the
    translation response of each segment won't exceed context length
    of the model. Also, run each segments currently also speeds up the job.
    :param groups_map: Dictionary of inline groups and their id
    :param max_token: max number of token in each segment
    :param token_counter: a function that takes a string as input and output number of tokens
    :return: A list of segmented inline groups
    """
    if not (token_all := sum([token_counter(str(g)) for g in groups_map.values()])):
        return []
    n_seg = math.ceil(token_all / max_token)
    len_seg = math.ceil(token_all / n_seg)
    ret = []
    token_cnt = 0
    cnt = 0
    seg = OrderedDict({})
    for k, group in groups_map.items():
        n = token_counter(str(group))
        if n > max_token:
            # raise ValueError(f'Length of single paragraph [{n}] exceed max length [{max_token}].')
            print(f'Single paragraph exceed max length [{n} > {max_token}]. Skip this one!')
            continue
        if (token_cnt > len_seg) and seg:
            ret.append(seg)
            token_cnt = 0
            cnt = 0
            seg = OrderedDict({})
        seg[str(cnt)] = group
        cnt += 1
        token_cnt += n
    ret.append(seg)
    return ret


async def translate_groups(
        groups_map: OrderedDict[str, InlineGroup],
        source_lang, target_lang, mapping_table, software_type, source_type, image_path=None, database_path=None, review_report_path=None
):
    """
    Performs translation and restruct task on one segment of all inline groups.
    :param groups_map: OrderedDict of inline groups and corresponding id
    :param source_lang: Source language
    :param target_lang: Target language
    :param mapping_table: Dictionary with specific name translations
    :param software_type: Type of software being translated
    :param source_type: Type of source file (e.g., 'UI', 'Help', etc.)
    :param image_path: Optional path to folder with images for translation enhancement
    :param database_path: Optional path to the database for searching similar pairs
    :return: list of results of each task,
             "S" for success;
             "C" for Compromise, which means cannot perfectly restruct the paragraph;
             "F" for Fail;
             or other exceptions
    """


    if image_path:
        print(f"Using images from {image_path} for translation enhancement")

    groups_in = {
        k: str(v).replace('\n', '') for k, v in groups_map.items()
    }
    # groups_in_str = json.dumps(groups_in, indent=0, ensure_ascii=False)
    groups_out = {}
    for source_text_index, source_text in groups_in.items():
        # Identify specific named entities in the text to translate
        relevant_specific_names = get_relevant_specific_names(mapping_table, source_text)
        print(f"Relevant specific names for translation: {relevant_specific_names}")

        # Search for relevant translated pairs in the database
        relevant_pair_database = []
        if database_path:
            relevant_pair_database = search_similar_pair_main(translate_dict={source_text_index: source_text}, database_path=database_path, grammar_top_n=5, term_top_n=5)
        print(f"Relevant specific names for translation: {relevant_pair_database}")
        
        # Initialize the chat with image_path if provided
        chat = OpenaiAPIChat(
            model_name=conf.TRANSLATE_MODEL,
            system_prompt=translate_sys_prompt(source_lang, target_lang, software_type, source_type),
            image_path=image_path
        )

        # print("===========================Used System Prompt=============================")
        # print(f"{chat.sys_prompt}")
        # print("===========================Used System Prompt=============================")

        response, stop_reason = '', ''
        try:
            p = translate_prompt(
                source_lang, 
                target_lang, 
                source_text,
                specific_names=relevant_specific_names,
                refer_data_list=relevant_pair_database,
            )
            async for chunk, stop_reason in chat.get_stream_aresponse(p, temperature=0.01):
                response += chunk
            
            if stop_reason == 'length':
                raise RuntimeError
        except RuntimeError:
            raise RuntimeError("Translation response exceeded length limit.")
        # print("===========================Used Prompt=============================")
        # print(f"{p}")
        # print("===========================Used Prompt=============================")
        print(f"Translation response:\n {response}")
        if not as_json_obj(response):
            print("Translation response is empty, breaking the loop.")
            continue        
        translated_text = list(as_json_obj(response).values())[-1]
        # Add await to properly call the async function
        translated_text, review_pass_flag = await review_n_improve_process(source_lang,
                                            target_lang,
                                            software_type,
                                            source_type,
                                            source_text, 
                                            translated_text, 
                                            relevant_specific_names,
                                            relevant_pair_database,
                                            image_path,
                                            model_list=conf.COMPARISON_MODEL, 
                                            temperature=conf.TEMPERATURE, 
                                            seed=conf.SEED,
                                            review_path=review_report_path)

        groups_out[source_text_index] = translated_text

        debug_process(source_text_index, source_text, relevant_specific_names, relevant_pair_database, p, response, list(as_json_obj(response).values())[-1])

    print(f"Translation response--2: {groups_out}")
    # Ensure groups_out has the exact same keys as groups_in to preserve structure
    if groups_out and set(groups_out.keys()) != set(groups_in.keys()):
        print("Warning: Translated structure doesn't match original structure. Attempting to fix...")
        fixed_groups_out = {}
        for key in groups_in.keys():
            if key in groups_out:
                fixed_groups_out[key] = groups_out[key]
            else:
                # If a key is missing in the translation, keep the original (untranslated)
                fixed_groups_out[key] = groups_in[key]
                print(f"Warning: Missing translation for group {key}, keeping original")
        
        # Check for any extra keys in groups_out that shouldn't be there
        extra_keys = set(groups_out.keys()) - set(groups_in.keys())
        if extra_keys:
            print(f"Warning: Found extra keys in translation response: {extra_keys}")
        
        groups_out = fixed_groups_out    
    # Check if the groups_map contains actual HTML elements or is from Excel (with None elements)
    is_excel_translation = all(group.elements[0] is None for group in groups_map.values())
    
    # Add await to properly call the async function
    await restruct_process(is_excel_translation, groups_in, groups_out, groups_map)

async def translation_pipeline(
        soup: BeautifulSoup,
        source_lang, target_lang, mapping_table, software_type, source_type, image_path=None, database_path=None, review_report_path=None
) -> list[str]:
    """
    Main entry point to translates the HTML in the soup object in place.
    The pipeline is composed of 3 steps:
    1. Traverse the DOM, groups up adjacent inline elements into InlineGroup,
        since, conventionally, these elements are more likely to form a semantically
        complete sentence or paragraph, thus makes the model more likely to generate
        a better translation.
    2. Split the inline groups to be translated into smaller translation tasks, so that
        the response won't exceed model's context length, then, run each tasks concurrently.
    3. After a translation tasks are done, use the model to split each translated
        inline group and fit each piece back to inline tags to recover the structure
        of the DOM.

    :param soup: The BeautifulSoup object representing the HTML to be translated
    :param source_lang: Source language
    :param target_lang: Target language
    :param mapping_table: Dictionary of specific name translations
    :param software_type: Type of software being translated
    :param source_type: Type of source file (e.g., 'UI', 'Help', etc.)
    :param image_path: Optional path to folder with images to enhance translation quality
    :param database_path: Optional path to the database for searching similar pairs
    :return: list of results of translation segments,
             "S" for success;
             "C" for Compromise, which means cannot perfectly restruct the paragraph;
             "F" for Fail, or any other exceptions
    """
    groups_map = get_text_group_inline(soup)
    groups_map_segments = segment_groups_map(
        groups_map,
        int(conf.N_INPUT_TOKEN),
        OpenaiAPIChat(conf.TRANSLATE_MODEL).n_tokens
    )
    tasks = [translate_groups(seg, source_lang, target_lang, mapping_table, software_type, source_type, image_path, database_path, review_report_path) for seg in groups_map_segments]
    results = await asyncio.gather(*tasks)
    results = [j for i in results for j in i]
    return results


def detect_file_type(content, file_path=""):
    """
    Detects if the given content is likely XML or HTML, with special handling for POMO XML and XLSX.
    First checks file extension, then falls back to content analysis.
    
    :param content: The file content as string
    :param file_path: The file path, used to check for file extension and POMO files by name
    :return: A tuple of (parser_type, is_pomo_xml, is_xlsx)
        parser_type: 'xml' if content appears to be XML, 'html.parser' for HTML, 'xlsx' for Excel
        is_pomo_xml: Boolean indicating if this is a POMO XML file
        is_xlsx: Boolean indicating if this is an Excel file
    """
    is_pomo_xml = False
    is_xlsx = False
    
    # First check by file extension if file_path is provided
    if file_path:
        file_extension = os.path.splitext(file_path.lower())[1]
        
        # Check for Excel files first
        if file_extension in ['.xlsx', '.xls']:
            return 'xlsx', False, True
            
        # Direct extension detection
        if file_extension in ['.xml', '.xaml']:
            # Check for POMO indicators in filename
            if 'pomo' in file_path.lower():
                is_pomo_xml = True
            return 'xml', is_pomo_xml, False
            
        elif file_extension in ['.html', '.htm']:
            return 'html', False, False
    
    # If file extension check doesn't determine the type, fallback to content analysis
    
    # Check filename for POMO indicators
    if file_path and ('pomo' in file_path.lower()):
        is_pomo_xml = True
    
    # Check for specific POMO XML structure patterns
    if re.search(r'<Entry\s+Id=|<Group\s+Id=', content):
        is_pomo_xml = True
    
    # Check for XML declaration
    if re.search(r'^\s*<\?xml', content):
        return 'xml', is_pomo_xml, False
    
    # Look for common HTML indicators
    if re.search(r'<!DOCTYPE\s+html|<html\b|<body\b|<head\b', content, re.IGNORECASE):
        return 'html', False, False
    
    # Check if the content follows XML structure patterns
    # XML typically has a single root element
    clean_content = re.sub(r'<!--.*?-->', '', content, flags=re.DOTALL)  # Remove comments
    clean_content = re.sub(r'<\?.*?\?>', '', clean_content)  # Remove processing instructions
    root_elements = re.findall(r'<([^\s/>]+)[^>]*>(?:.*?)</\1>', clean_content, re.DOTALL)
    if len(root_elements) == 1 or is_pomo_xml:
        return 'xml', is_pomo_xml, False
    
    # Default to HTML for safety
    return 'html', False, False


async def translate_xlsx(
        input_file, 
        output_file, 
        source_lang, 
        target_langs, 
        mapping_table, 
        software_type, 
        source_type, 
        image_path=None, 
        database_path=None,
        review_report_path=None):
    """
    Translate an Excel file where the first column contains text to be translated.
    Sends all text to API at once as JSON for more efficient translation.
    All target languages will be added as columns in the same output file.
    
    :param input_file: Path to input XLSX file
    :param output_file: Path to output XLSX file
    :param source_lang: Source language
    :param target_langs: List of target languages
    :param mapping_table: Dictionary with specific name translations
    :param software_type: Type of software being translated
    :param source_type: Type of source file (e.g., 'UI', 'Help', etc.)
    :param image_path: Optional path to images for translation enhancement
    :param database_path: Optional path to database for translation
    :return: Dictionary with success/error counts
    """
    print(f"Starting XLSX translation from {input_file} to {output_file}")
    print(f"Source language: {source_lang}, Target languages: {', '.join(target_langs)}")
    
    try:
        # Read the Excel file with pandas
        df = pd.read_excel(input_file)
        
        # Check if the file has data
        if df.empty:
            print("Error: Input Excel file is empty")
            return {"success": 0, "error": 1}
        
        # Get the name of the first column, which should contain source language text
        source_column = df.columns[0]
        print(f"Found input column: {source_column}")
        
        # Create a single output dataframe that will contain all translations
        output_df = df.copy()
        
        # Process each target language and add as a new column
        for lang in target_langs:
            print(f"Translating to {lang}...")
            
            # For multi-language options, use specific mapping table for each language if available
            current_mapping = mapping_table

              # Extract all text to translate into a dictionary
            text_to_translate = {}
            for i, row in df.iterrows():
                # Skip empty cells or non-string values
                if pd.isna(row[source_column]) or not isinstance(row[source_column], str):
                    continue
                
                text_value = str(row[source_column])
                text_to_translate[str(i)] = text_value
            
            # Create an OrderedDict that mimics the structure expected by translate_groups
            groups_map = OrderedDict({})
            for idx, text in text_to_translate.items():
                # Create a dummy InlineGroup for each text entry
                # This allows us to use the translate_groups function
                text_list = [text]
                cids_list = [0]  # Dummy cid
                elements_list = [None]  # We don't need actual elements for this use case
                groups_map[idx] = InlineGroup(text_list, cids_list, elements_list)
            
            # Segment the groups if needed to respect token limits
            groups_map_segments = segment_groups_map(
                groups_map,
                int(conf.N_INPUT_TOKEN),
                OpenaiAPIChat(conf.TRANSLATE_MODEL).n_tokens
            )
            
            print(f"Split the text into {len(groups_map_segments)} segments for translation")
            
            # Process each segment and gather results
            all_translated_results = {}
            for segment in groups_map_segments:
                print(f"Processing segment with {len(segment)} text entries...")                
                # Use the existing translate_groups function
                responses = await translate_groups(
                    segment, 
                    source_lang, 
                    lang, 
                    current_mapping, 
                    software_type, 
                    source_type,
                    image_path,
                    database_path,
                    review_report_path
                )
                
                # Process the translated texts
                for i, group_key in enumerate(segment.keys()):
                    try:
                        # Get the translated text from the response
                        if isinstance(responses[i], str):
                            all_translated_results[group_key] = responses[i]
                        else:
                            # Handle case where response might be a success code
                            print(f"Warning: Unexpected response type for item {group_key}")
                            all_translated_results[group_key] = text_to_translate[group_key]
                    except Exception as e:
                        print(f"Warning: Error processing translation for item {group_key}: {e}")
                        all_translated_results[group_key] = text_to_translate[group_key]
            
            # Add translated text back to the dataframe
            lang_column_name = conf.LANGUAGE_MAP.get(lang, lang)
            for idx_str, translated_text in all_translated_results.items():
                try:
                    idx = int(idx_str)
                    # Only add if the index exists in the dataframe
                    if idx < len(output_df):
                        output_df.loc[idx, lang_column_name] = translated_text
                except (ValueError, KeyError) as e:
                    print(f"Error adding translation at index {idx_str}: {e}")
                    
            print(f"Added translations for {lang} as column '{lang_column_name}'")
                # Save the output dataframe to Excel
        print(f"Saving Excel file with {len(output_df)} rows and {len(output_df.columns)} columns")
        
        # Use openpyxl to save with nice formatting
        # First save with pandas
        output_df.to_excel(output_file, index=False)
        
        # Then apply formatting with openpyxl
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Format header row
        for col_idx, column in enumerate(output_df.columns):
            # Bold header
            cell = ws.cell(row=1, column=col_idx+1)
            cell.font = openpyxl.styles.Font(bold=True)
            
            # Auto-adjust column width based on content length
            max_length = len(str(column))
            for i, value in enumerate(output_df[column]):
                if value:
                    max_length = max(max_length, len(str(value)))
                    
            adjusted_width = min(max(max_length + 2, 10), 80)  # Min 10, Max 80
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx+1)].width = adjusted_width
        
        # Save the formatted workbook
        wb.save(output_file)
        
        print(f"Excel translation completed. Output saved to {output_file} with {len(target_langs)} language columns.")
        return {"success": 1, "error": 0}
    
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        import traceback
        traceback.print_exc()
        return {"success": 0, "error": 1}


def main(p_in="default", 
         p_out="default", 
         source_lang="default", 
         target_lang="default", 
         specific_names_xlsx="default", 
         software_type="default", 
         image_path="default",
         source_type="default",
         database_path="default",
         review_report_path="default"):
    # Run in translation mode
    print("Running in translation mode...")
    if p_in=="default":
        p_in = conf.INPUT_FILE_PATH
    if p_out=="default":
        p_out = conf.OUTPUT_FILE_PATH

    if source_lang=="default":
        source_lang = conf.SOURCE_LANGUAGE
    
    if target_lang=="default":
        target_lang = conf.TARGET_LANGUAGE

    if specific_names_xlsx=="default":
        specific_names_xlsx = conf.SPECIFIC_NAMES_XLSX

    if software_type=="default":
        software_type = conf.SOFTWARE_TYPE
        
    if image_path=="default" and hasattr(conf, 'IMAGE_PATH'):
        image_path = conf.IMAGE_PATH

    if image_path == False: # This is a flag to disable image path for the request from batch_process.py
        print("Image path is set to False, no images will be used for translation enhancement.")
        image_path = "default"

    if source_type=="default":
        source_type = conf.SOURCE_TYPE

    if database_path=="default":
        database_path = conf.DATABASE_PATH

    if review_report_path == "default":
        review_report_path = conf.REVIEW_REPORT_PATH

    print(f'Image path: {image_path}')

    # Check if this is a multi-language request
    target_languages = []
    if target_lang in conf.MULTI_LANGUAGE_OPTIONS:
        print(f"Multi-language option '{target_lang}' detected. Will translate to {len(conf.MULTI_LANGUAGE_OPTIONS[target_lang])} languages.")
        target_languages = conf.MULTI_LANGUAGE_OPTIONS[target_lang]
        
        # For multi-language, we need to process each language
        for current_target in target_languages:
            # Create output path specific to the current language
            # Use the original path but add the language code to filename
            file_dir = os.path.dirname(p_out)
            file_name, file_ext = os.path.splitext(os.path.basename(p_out))
            current_output_path = os.path.join(file_dir, f"{file_name}_{current_target.replace(' ', '')}{file_ext}")
            print(f"\n--- Processing language: {current_target} ---")
            print(f"Output path: {current_output_path}")
            
            # Call process_single_file to handle this language
            process_single_file(p_in, current_output_path, source_lang, current_target, specific_names_xlsx, software_type, source_type, image_path, database_path, review_report_path)
        
        return
    
    # Single language processing
    process_single_file(p_in, p_out, source_lang, target_lang, specific_names_xlsx, software_type, source_type, image_path, database_path)


def process_single_file(p_in, p_out, source_lang, target_lang, specific_names_xlsx, software_type, source_type, image_path=None, database_path=None, review_report_path=None):
    """Process a single file translation"""
    print(f"Input file: {p_in}")
    print(f"Output file: {p_out}")
    print(f"Source language: {source_lang}")
    print(f"Target language: {target_lang}")
    print(f"Specific names file: {specific_names_xlsx}")
    print(f"Software type: {software_type}")
    print(f"Source type: {source_type}")
    print(f"Image path: {image_path}" if image_path else "No image path provided")
    print(f"Database path: {database_path}" if database_path else "No database path provided")
    print(f"Review report path: {review_report_path}" if review_report_path else "No review report path provided")

    # Load specific names dictionary
    mapping_table = {}
    if specific_names_xlsx:
        mapping_table = load_specific_names(specific_names_xlsx, source_lang, target_lang)
        # print(f"Loaded specific names: {mapping_table}")

    try:
        # Check if the file exists
        if not os.path.exists(p_in):
            raise FileNotFoundError(f"Input file not found: {p_in}")
            
        # Try to get language code from config and read the file
        used_encoding, file_content = detect_file_encoding(p_in, source_lang)
        print(f"Using {used_encoding} encoding for input file")
    except Exception as e:
        print(f"ERROR: {e}")
        return
    
    # Detect if the input is XML, HTML or XLSX
    file_type, is_pomo_xml, is_xlsx = detect_file_type(file_content, p_in)
    print(f"Detected file type: {file_type}, Is POMO XML: {is_pomo_xml}, Is XLSX: {is_xlsx}")

    # Store the original content for POMO XML files to preserve exact tag case
    original_content = file_content if is_pomo_xml else None

    # Handle XLSX input files
    if is_xlsx or file_type == 'xlsx':
        print("Processing XLSX file...")
        
        # Check if the target language is a multi-language option
        if target_lang in conf.MULTI_LANGUAGE_OPTIONS:
            target_languages = conf.MULTI_LANGUAGE_OPTIONS[target_lang]
            print(f"Multi-language option '{target_lang}' selected. Translating to: {', '.join(target_languages)}")
        else:
            # Single language case
            target_languages = [target_lang]
            print(f"Single language translation to: {target_lang}")
            
        # For XLSX files, we run special translation procedure
        result = asyncio.run(translate_xlsx(p_in, p_out, source_lang, target_languages, mapping_table, software_type, source_type, image_path, database_path, review_report_path))
        
        # Check the result
        if result["success"]:
            print(f"XLSX translation completed successfully. Output saved to {p_out}")
        else:
            print(f"XLSX translation encountered errors.")
        return

    # Handle HTML files    
    if file_type == 'html':
        bs = BeautifulSoup(file_content, 'html.parser')
        ret = asyncio.run(translation_pipeline(bs, source_lang, target_lang, mapping_table, software_type, source_type, image_path, database_path, review_report_path))
        
        # Use the same encoding for writing
        with open(p_out, 'w', encoding=used_encoding) as fout:
            fout.write(str(bs))

    # Handle XML files
    else:
        print('Start to translate XML...')
        bs = BeautifulSoup(file_content, file_type)
        ret = asyncio.run(translation_pipeline(bs, source_lang, target_lang, mapping_table, software_type, source_type, image_path, database_path, review_report_path))

        # Use the same encoding for writing
        with open(p_out, 'w', encoding=used_encoding) as fout:
            # For POMO XML files, we'll use a regex-based approach to preserve the exact case of tags
            if is_pomo_xml:
                # Get the translated content
                translated_content = str(bs)
                
                # Create a mapping of original tag formats
                tag_patterns = {}
                # Find all opening tags with their attributes in the original content
                for match in re.finditer(r'<([A-Za-z]+)(\s+[^>]*)?>', original_content):
                    full_tag = match.group(0)
                    tag_name = match.group(1)
                    tag_patterns[tag_name.lower()] = tag_name
                
                # Replace tags in translated content with original case
                for lower_tag, original_tag in tag_patterns.items():
                    # Replace opening tags
                    translated_content = re.sub(
                        r'<(' + lower_tag + r')(\s+[^>]*?)>', 
                        lambda m: '<' + original_tag + m.group(2) + '>', 
                        translated_content
                    )
                    # Replace closing tags
                    translated_content = re.sub(
                        r'</(' + lower_tag + r')>', 
                        lambda m: '</' + original_tag + '>', 
                        translated_content
                    )
                
                output_content = translated_content
            else:
                # Normal output for HTML or regular XML
                output_content = str(bs)
            
            output_content = re.sub(
                r'<\?xml\s+version="[^"]*"\s+encoding="[^"]*"\s*\?>\s*',
                '',
                output_content
            )

            fout.write(output_content)
            
        print(f"Translation completed: {ret.count('S')} successful, {ret.count('C')} compromised, {ret.count('F')} failed out of {len(ret)} segments")
    
    print(f"Output file written using {used_encoding} encoding")

if __name__ == '__main__':
    main()
