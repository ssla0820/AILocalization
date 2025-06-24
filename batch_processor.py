"""
Batch processor for HTML/XML translations using Excel configuration file.
This script processes an Excel file row by row, applying translation, verification,
and ground truth checking for each configuration row.

The batch Excel file can include the following columns:
- Required: INPUT_FILE_PATH_FOLDER, OUTPUT_FILE_PATH_FOLDER, COMPARE_FILE_PATH_FOLDER, 
  SOURCE_LANGUAGE, TARGET_LANGUAGE, SOFTWARE_TYPE
- Optional: SPECIFIC_NAMES_XLSX_PATH, IMAGE_PATH_FOLDER, CHECK_VERIFICATION, CHECK_GROUND_TRUTH, GROUND_TRUTH_PATH

Verification compares the source and translated files to check for translation quality.
To skip verification for a specific row, set CHECK_VERIFICATION column to "False".

Ground truth checking verifies translations against predefined expected values in a ground truth Excel file.
To enable ground truth checking, set CHECK_GROUND_TRUTH column to "True" for desired rows.
You can optionally specify a custom GROUND_TRUTH_PATH for each row.
"""

import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__))))
# import asyncio
import pandas as pd
from translate.translate import main as translate_main
from interface.run_interface import run_translation_interface
# from verify import main as verify_main
from groundtruth_check.GroundTruth_Check import main as groundtruth_main
from config import translate_config
# import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from config import translate_config as conf



def ensure_dir(dir_path):
    """Create directory if it doesn't exist"""
    if not os.path.exists(dir_path):
        os.makedirs(dir_path, exist_ok=True)

def get_files_to_process(input_folder, extensions=['.html', '.htm', '.xml', '.xlsx', '.xls']):

    """
    Get list of files to process from the input folder.
    
    :param input_folder: Folder containing files to translate
    :param extensions: List of file extensions to include
    :return: List of file paths
    """
    files = []
    for root, _, filenames in os.walk(input_folder):
        for filename in filenames:
            _, ext = os.path.splitext(filename)
            if ext.lower() in extensions:
                full_path = os.path.join(root, filename)
                files.append(full_path)
    return files

def get_output_filename(input_path, output_folder, target_language, is_compare_file=False, is_ground_truth_file=False, is_review_file=False):
    """
    Generate output filename with target language suffix.
    
    :param input_path: Input file path
    :param output_folder: Output folder
    :param target_language: Target language for translation
    :param is_compare_file: Whether this is a comparison file (always use .html extension)
    :return: Output file path
    """
    filename = os.path.basename(input_path)
    name, ext = os.path.splitext(filename)
    
    # Create language suffix by removing spaces
    lang_suffix = target_language.replace(' ', '')
    
    # Handle file name conflicts with 2-digit counter (01, 02, etc.)
    counter = 1

    # For comparison files, always use .html extension
    if is_compare_file:
        ext = ".html"
    elif is_ground_truth_file or is_review_file:
        ext = ".xlsx"
    
    if is_review_file:
        output_filename = f"{name}_{lang_suffix}_Review_{counter:02d}{ext}"
        output_path = os.path.join(output_folder, output_filename)
    else:
        # Generate output filename
        output_filename = f"{name}_{lang_suffix}_{counter:02d}{ext}"
        output_path = os.path.join(output_folder, output_filename)
        

    while os.path.exists(output_path):
        output_filename = f"{name}_{lang_suffix}_{counter:02d}{ext}"
        output_path = os.path.join(output_folder, output_filename)
        counter += 1
        
    if is_compare_file or is_ground_truth_file or is_review_file:
        return output_path
    
    return output_path, f"{name}_{lang_suffix}"

def get_refer_text_n_image_path(input_path, refer_text_table_folder):
    filename = os.path.basename(input_path)
    name, ext = os.path.splitext(filename)

    return os.path.join(refer_text_table_folder, f"{name}_refer_text_n_image.xlsx")

def get_multi_language_xlsx_output(input_path, output_folder, target_languages, multi_language_code=None):
    """
    Generate output filename for multi-language Excel translations in a single file.
    
    :param input_path: Input file path
    :param output_folder: Output folder
    :param target_languages: List of target languages
    :param multi_language_code: Original multi-language code (e.g. '2L', '9L')
    :return: Output file path
    """
    filename = os.path.basename(input_path)
    name, ext = os.path.splitext(filename)
    
    # Use the original multi-language code if provided, otherwise create a descriptive suffix
    if multi_language_code:
        lang_suffix = multi_language_code
    else:
        # For just two languages, show both names - for more than two, show count
        if len(target_languages) <= 2:
            lang_suffix = "_".join([lang.replace(' ', '') for lang in target_languages])
        else:
            lang_suffix = f"MultiLang_{len(target_languages)}"
    
    # Handle file name conflicts with 2-digit counter (01, 02, etc.)
    counter = 1
    
    # Generate output filename
    output_filename = f"{name}_{lang_suffix}_{counter:02d}{ext}"
    output_path = os.path.join(output_folder, output_filename)
    
    while os.path.exists(output_path):
        counter += 1
        output_filename = f"{name}_{lang_suffix}_{counter:02d}{ext}"
        output_path = os.path.join(output_folder, output_filename)
    
    return output_path

def create_results_excel(results_file_path):
    """
    Creates or updates the results Excel file with appropriate headers.
    
    :param results_file_path: Path to the results Excel file
    :return: None
    """
    # Check if file exists
    if not os.path.exists(results_file_path):
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Translation Results"
          # Add headers
        headers = [
            "Source Path", 
            "Output Path", 
            "Source Language", 
            "Target Language", 
            "Translation Status", 
            "Failed Sentences"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Save the file
        wb.save(results_file_path)
        print(f"Created new results file: {results_file_path}")

def add_result_to_excel(results_file_path, result_data):
    """
    Adds a result entry to the results Excel file.
    
    :param results_file_path: Path to the results Excel file
    :param result_data: Dictionary containing result data
    :return: None
    """
    # Create the file if it doesn't exist
    if not os.path.exists(results_file_path):
        create_results_excel(results_file_path)
    
    # Load the workbook
    wb = openpyxl.load_workbook(results_file_path)
    ws = wb.active
    
    # Get the next row
    next_row = ws.max_row + 1
    
    # Add data
    ws.cell(row=next_row, column=1).value = result_data.get('source_path', '')
    ws.cell(row=next_row, column=2).value = result_data.get('output_path', '')
    ws.cell(row=next_row, column=3).value = result_data.get('source_language', '')
    ws.cell(row=next_row, column=4).value = result_data.get('target_language', '')
    
    # Translation status with color coding
    translation_cell = ws.cell(row=next_row, column=5)
    translation_cell.value = result_data.get('translation_status', 'Failed')
    if translation_cell.value == 'Success':
        translation_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        translation_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")    # Verification status with color coding

    # Failed sentences
    ws.cell(row=next_row, column=6).value = result_data.get('failed_sentences', '')
    
    # Save the workbook
    wb.save(results_file_path)

def extract_failed_sentences(compare_file_path):
    """
    Extracts failed sentences from comparison HTML file.
    
    :param compare_file_path: Path to the comparison HTML file
    :return: String of failed sentences, or empty string if none found or file doesn't exist
    """
    if not os.path.exists(compare_file_path):
        return ""
    
    try:
        from bs4 import BeautifulSoup
        
        # Try different encodings
        encodings = ['utf-8', 'utf-16', 'iso-8859-1', 'cp1252']
        content = None
        
        for encoding in encodings:
            try:
                with open(compare_file_path, 'r', encoding=encoding) as file:
                    content = file.read()
                break
            except UnicodeDecodeError:
                continue
        
        if content is None:
            print(f"Could not read comparison file with available encodings: {compare_file_path}")
            return ""
        
        # Parse HTML
        soup = BeautifulSoup(content, 'html.parser')
        
        # Look for elements with red background which typically indicate issues
        problem_elements = soup.find_all(style=lambda s: s and 'background-color: #ffcccc' in s)
        
        if not problem_elements:
            # Try different approach - find table rows with issues
            problem_rows = soup.find_all('tr', class_=lambda c: c and 'issue' in c.lower() if c else False)
            problem_elements = []
            for row in problem_rows:
                cells = row.find_all('td')
                if cells and len(cells) >= 2:
                    problem_elements.append(cells[1])  # Assuming second cell contains the translation
        
        failed_sentences = [elem.get_text().strip() for elem in problem_elements]
        return "; ".join(failed_sentences)
    
    except Exception as e:
        print(f"Error extracting failed sentences from {compare_file_path}: {str(e)}")
        return ""

def merge_xlsx_translations(original_file, translated_files, output_file):
    """
    Merge multiple translated Excel files into a single file with all languages as columns
    
    :param original_file: Path to the original input Excel file
    :param translated_files: List of translated Excel files to merge
    :param output_file: Path to save the merged output file
    :return: True if successful, False otherwise
    """
    try:
        print(f"Merging {len(translated_files)} translated XLSX files into: {output_file}")
        
        # Read the original file to get the base structure
        original_df = pd.read_excel(original_file)
        
        # Create a new dataframe starting with the original content
        merged_df = original_df.copy()
        
        # Process each translated file
        for trans_file in translated_files:
            try:
                # Get the language from the filename
                filename = os.path.basename(trans_file)
                name_parts = os.path.splitext(filename)[0].split('_')
                
                # The language should be in the second-to-last part (before the counter)
                if len(name_parts) < 3:
                    print(f"Warning: Could not determine language from filename: {filename}")
                    lang_name = f"Translation_{len(merged_df.columns) + 1}"
                else:
                    lang_name = name_parts[-2]
                
                # Read the translated file
                trans_df = pd.read_excel(trans_file)
                
                # Check if the dataframe has at least 2 columns (source + translation)
                if len(trans_df.columns) < 2:
                    print(f"Warning: Translated file has insufficient columns: {trans_file}")
                    continue
                
                # Get the second column (translation column)
                trans_column = trans_df.columns[1]
                
                # Add this column to the merged dataframe
                merged_df[lang_name] = trans_df[trans_column]
                
                print(f"Added {lang_name} translations from {trans_file}")
                
            except Exception as e:
                print(f"Error processing translated file {trans_file}: {str(e)}")
        
        # Save the merged dataframe
        merged_df.to_excel(output_file, index=False)
        
        # Apply formatting using openpyxl
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Format the header row
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            
            # Auto-fit column width based on content
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max(max_length + 2, 10), 80)  # Min 10, Max 80
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = adjusted_width
        
        wb.save(output_file)
        print(f"Successfully created merged Excel file: {output_file}")
        return True
        
    except Exception as e:
        print(f"Error merging Excel files: {str(e)}")
        return False

def process_batch_file(task):
    """
    Process the batch file containing translation tasks.
    This function reads the batch Excel file, processes each row,
    and performs translations, verifications, and ground truth checks as specified.
    :param tasks: List of translation task configurations
    :return: None
    """

    # try:
    success_count = 0
    error_count = 0
    results_files = []
    

    source_type = task.get('source_type')
    product_name = task.get('product_name')
    software_type = conf.SOFTWARE_TYPE_MAP.get(product_name, '')
    source_language = task.get('source_lang')
    target_language = task.get('target_lang')
    input_folder = task.get('input_folder')
    output_folder = task.get('output_folder')
    review_folder = task.get('review_folder')
    

    # Check if this is a multi-language option
    target_languages = [target_language]
    is_multi_language = False
    if target_language in conf.MULTI_LANGUAGE_OPTIONS:
        print(f"Multi-language option '{target_language}' detected. Will translate to {len(conf.MULTI_LANGUAGE_OPTIONS[target_language])} languages.")
        target_languages = conf.MULTI_LANGUAGE_OPTIONS[target_language]
        is_multi_language = True
    
    specific_names_xlsx = task.get('glossary_file_path') if task.get('glossary_file_path') != 'None' else None
    refer_text_table_folder = task.get('refer_info') if task.get('refer_info') != 'None' else None

    if task.get('translation_memory')['folder'] == 'None':
        database_path_list = None
    else:
        database_path_list = [task.get('translation_memory').get('files')]
    
    if task.get('common_term')['folder'] == 'None':
        region_table_path_list = None
    else:
        region_table_path_list = [task.get('common_term').get('files')]

    
    
    print(f"Configuration: {source_language} -> {', '.join(target_languages)}, Software: {software_type}")
    print(f"Source type: {source_type}")
    print(f"Input folder: {input_folder}")
    print(f"Output folder: {output_folder}")
    print(f"Review folder: {review_folder}")

    # Create output directories if they don't exist
    ensure_dir(output_folder)
    ensure_dir(review_folder)

    results_file = os.path.join(output_folder, f"translation_results_{source_language}_to_{target_language}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    create_results_excel(results_file)
    results_files.append(results_file)
    
    # For tracking XLSX files translations that need merging
    xlsx_files_to_merge = {}  # input_file -> list of translated files
    

    for lang_index in range(len(target_languages)):
        current_target_language = target_languages[lang_index]

        if database_path_list is None: database_path = None
        else: database_path = os.path.join(task.get('translation_memory')['folder'], database_path_list[lang_index][current_target_language])

        if region_table_path_list is None: region_table_path = None
        else: region_table_path = os.path.join(task.get('common_term')['folder'], region_table_path_list[lang_index][current_target_language])


        print(f"\n--- Processing language: {current_target_language} ---")
        # Get all files to process from input folder
        files_to_process = get_files_to_process(input_folder)
        print(f"Found {len(files_to_process)} files to process")
        
        if not files_to_process:
            print(f"No files found in input folder: {input_folder}")
            continue
        
        # Process each file
        for i, input_file in enumerate(files_to_process, 1):
            try:
                print('='*100)
                print(f"Processing file {i} of {len(files_to_process)}: {input_file}")
                print('='*100)

                # Check if this is an Excel file
                is_excel = input_file.lower().endswith(('.xlsx', '.xls'))
                
                # Generate output and compare file paths
                output_file, ground_truth_file_name = get_output_filename(input_file, output_folder, current_target_language)
                review_file = get_output_filename(input_file, review_folder, current_target_language, is_review_file=True)
                refer_text_table_path = get_refer_text_n_image_path(input_file, refer_text_table_folder)
                print(f"Output file: {output_file}")

                
                # Get the file-specific image path folder if available
                file_specific_image_path = False

                # Prepare result data dictionary
                result_data = {
                    'source_path': input_file,
                    'output_path': output_file,
                    'source_language': source_language,
                    'target_language': current_target_language,
                    'translation_status': 'Failed',
                    'verification_status': 'Failed',
                    'groundtruth_status': 'N/A',
                    'failed_sentences': ''
                }
                
                # Run translation
                print(f"Starting translation...")
                translation_success = False
                try:
                    translate_main(input_file, 
                                    output_file, 
                                    source_language, 
                                    current_target_language,
                                    specific_names_xlsx, 
                                    region_table_path, 
                                    refer_text_table_path, 
                                    software_type, 
                                    image_path=file_specific_image_path,
                                    source_type=source_type,
                                    database_path=database_path,
                                    review_report_path=review_file)
                    
                    translation_success = os.path.exists(output_file)
                    result_data['translation_status'] = 'Success' if translation_success else 'Failed'
                    print(f"Translation completed: {output_file}")
                    
                    # If this is an Excel file and it's part of a multi-language translation,
                    # add it to the list of files to merge later
                    if is_excel and is_multi_language and translation_success:
                        if input_file not in xlsx_files_to_merge:
                            xlsx_files_to_merge[input_file] = []
                        xlsx_files_to_merge[input_file].append(output_file)
                        print(f"Added {output_file} to Excel files to be merged later")
                    
                except Exception as e:
                    print(f"Error in translation: {str(e)}")
                    result_data['translation_status'] = 'Failed'

                # Add result to Excel file
                add_result_to_excel(results_file, result_data)
                
                # Track overall success/error counts
                process_success = translation_success

                if process_success:
                    success_count += 1
                else:
                    error_count += 1
            except Exception as e:
                print(f"Error processing file {input_file}: {str(e)}")
                
                # Add failed result to Excel
                result_data = {
                    'source_path': input_file,
                    'output_path': output_file if 'output_file' in locals() else '',
                    'source_language': source_language,
                    'target_language': current_target_language,
                    'translation_status': 'Failed',
                    'failed_sentences': str(e)
                }
                add_result_to_excel(results_file, result_data)
                
                error_count += 1
        
        # After processing all languages, merge the Excel files if needed
        if xlsx_files_to_merge:
            print(f"\n--- Merging Excel files for multi-language translations ---")
            
            for input_file, translated_files in xlsx_files_to_merge.items():
                # Only merge if we have multiple files
                if len(translated_files) > 1:
                    # Generate output filename for merged file with the original multi-language code
                    merged_output = get_multi_language_xlsx_output(input_file, output_folder, target_languages, target_language)
                    
                    print(f"Merging translations for {os.path.basename(input_file)} into {os.path.basename(merged_output)}")
                    
                    # Merge the files
                    merge_result = merge_xlsx_translations(input_file, translated_files, merged_output)
                    # Add result to Excel file
                    merge_result_data = {
                        'source_path': input_file,
                        'output_path': merged_output,
                        'source_language': source_language,
                        'target_language': target_language,  # Use the original multi-language code (e.g. '2L', '9L')
                        'translation_status': 'Success' if merge_result else 'Failed',
                        'failed_sentences': ''
                    }
                    add_result_to_excel(results_file, merge_result_data)
                    
                    if merge_result:
                        print(f"Successfully merged Excel translations into: {merged_output}")
                    else:
                        print(f"Failed to merge Excel translations")
        
    
    print(f"Batch processing completed. Success: {success_count}, Errors: {error_count}")
    print(f"Results saved to:")
    for results_file in results_files:
        print(f"  - {results_file}")
        
    return {"success": success_count, "error": error_count, "results_files": results_files}
        
    # except Exception as e:
    #     print(f"Error processing batch file: {str(e)}")
    #     return {"success": 0, "error": 1}

def main(batch_excel_path=None):
    """
    Main entry point for batch processing.
    
    :param batch_excel_path: Path to the batch Excel file
    """
    # tasks = run_translation_interface()
    tasks = [{'source_type': 'UI', 'product_name': 'PDR', 'source_lang': 'English', 'target_lang': '2L', 'input_folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Source', 'refer_info': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Refer', 'output_folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Output', 'review_folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Review', 'glossary_file_path': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Mapping_Table\\specific_name_pdr365_June_2.xlsx', 'translation_memory': {'folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v11\\database', 'files': {'Traditional Chinese': 'PDR_CHT_translation_memory.json', 'Simplified Chinese': 'PDR_CHS_translation_memory.json'}}, 'common_term': {'folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v11\\region_table', 'files': {'Traditional Chinese': 'PDR_CHT_common_table.xlsx', 'Simplified Chinese': 'PDR_CHS_common_table.xlsx'}}}]
    # tasks = [
    #     # {'source_type': 'UI', 
    #     #       'product_name': 'PDR', 
    #     #       'source_lang': 'English', 
    #     #       'target_lang': '2L', 
    #     #       'input_folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Source', 
    #     #       'refer_info': 'None', 
    #     #       'output_folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Output', 
    #     #       'review_folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Review', 
    #     #       'glossary_file_path': 'None', 'translation_memory': {'folder': 'None', 'files': {'default': 'None'}}, 
    #     #       'common_term': {'folder': 'None', 'files': {'default': 'None'}}}, 
              
    #           {'source_type': 'UI', 
    #            'product_name': 'PDR', 
    #            'source_lang': 'English', 
    #            'target_lang': 'German', 
    #            'input_folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Source', 
    #            'refer_info': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Refer', 
    #            'output_folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Output',
    #              'review_folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Batch_Test\\Review', 
    #              'glossary_file_path': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v8\\Report\\Mapping_Table\\specific_name_pdr365_June_2.xlsx', 
    #              'translation_memory': {'folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v11\\database', 'files': {'German': 'PDR_DEU_translation_memory.json'}}, 
    #              'common_term': {'folder': 'E:\\Debby\\9_Scripts\\TranslateHTML\\Translate_HTML_XML_v11\\region_table', 'files': {'German': 'PDR_DEU_common_table.xlsx'}}}]

    # if tasks:
    #     print(f"Received {len(tasks)} translation tasks from the interface:")
    #     for i, task in enumerate(tasks):
    #         print(f"\nTask {i+1}:")
    #         print(f"  Source Type: {task.get('source_type')}")
    #         print(f"  Product Name: {task.get('product_name')}")
    #         print(f"  Source Language: {task.get('source_lang')}")
    #         print(f"  Target Language: {task.get('target_lang')}")
    #         print(f"  Input File Path: {task.get('input_file_path')}")
    for task in tasks:
        print(f"\nProcessing task: {task}")
        results = process_batch_file(task)
    
    print(f"Batch processing summary:")
    print(f"  Success: {results['success']} files")
    print(f"  Errors: {results['error']} files")
    
    if 'results_files' in results:
        print(f"  Results saved to:")
        for results_file in results['results_files']:
            print(f"    - {results_file}")
    
    return results

if __name__ == "__main__":
    main()