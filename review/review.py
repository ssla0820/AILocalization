"""
Translation review module to compare two HTML/XML/Excel translations.
This module provides functionality to compare two files with translations
and generate an HTML report with the review results.
"""
import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir)))
from bs4 import BeautifulSoup
import os
import asyncio
import re
import pandas as pd
from chat.openai_api_chat import OpenaiAPIChat
from chat.gemini_api_chat import GeminiAPIChat
from prompts.review_prompts import *
from prompts.improve_prompts import *
from config import translate_config as conf
from pages.general_functions import *
from database.search_similar_pair import main as search_similar_pair_main
import json

def export_debug_info(
    source_text,
    translated_text,
    relevant_specific_names,
    relevant_region_table,
    relevant_refer_text_table,
    relevant_pair_database,
    review_prompt_obj,
    raw_review_response_dict,
    review_result_dict,
    model_name,
    output_file="debug_review.xlsx"
):
    """
    Export debugging information to an Excel file.
    
    Args:
        source_text: The source text being reviewed
        translated_text: The translated text being reviewed
        relevant_specific_names: Dictionary of specific names relevant to this text
        relevant_region_table: Region table information
        relevant_refer_text_table: Reference text table
        relevant_pair_database: Similar pairs from the database
        review_prompt_obj: The ReviewPrompts object with all prompts
        raw_review_response_dict: Raw responses from all review models
        review_result_dict: Final review result dictionary
        model_name: The name of the model used
        output_file: Output file path
        
    Returns:
        None
    """
    debug_info = {}

    
    # Basic information
    debug_info["Model Name"] = model_name
    debug_info["Source Text"] = source_text
    debug_info["Translated Text"] = translated_text
    debug_info["Final Translated Text"] = review_result_dict.get("final_translated_text", "")
    
    # Reference data
    debug_info["Specific Names"] = json.dumps(relevant_specific_names, ensure_ascii=False, indent=2) if relevant_specific_names else "{}"
    debug_info["Region Table"] = json.dumps(relevant_region_table, ensure_ascii=False, indent=2) if relevant_region_table else "{}"
    debug_info["Refer Text Table"] = json.dumps(relevant_refer_text_table, ensure_ascii=False, indent=2) if relevant_refer_text_table else "{}"
    debug_info["Similar Pairs"] = json.dumps(relevant_pair_database, ensure_ascii=False, indent=2) if relevant_pair_database else "[]"
    
    # System prompts for all 6 rooms
    debug_info["System Prompt"] = review_prompt_obj.sys_prompt
    
    # Prompts for all 6 rooms
    debug_info["Prompt"] = review_prompt_obj.review_prompt()
    
    # Responses from the model
    if raw_review_response_dict:
        for category, response in raw_review_response_dict.items():
            debug_info[f"Response - {category}"] = json.dumps(response, ensure_ascii=False, indent=2) if response else "None"
    
    # Review result data
    for key, value in review_result_dict.items():
        if key.startswith(f"{model_name}_review_"):
            debug_info[f"Result - {key}"] = value
    
    # Convert debug info to DataFrame
    df = pd.DataFrame([debug_info])
    
    # Save to Excel with proper encoding for East Asian languages
    if os.path.exists(output_file):
        try:
            existing_df = pd.read_excel(output_file, engine='openpyxl')
            final_df = pd.concat([existing_df, df], ignore_index=True)
        except Exception as e:
            print(f"Error reading existing file: {e}")
            final_df = df
    else:
        final_df = df
    
    # Save to Excel with openpyxl engine for better East Asian character support
    final_df.to_excel(output_file, index=False, engine='openpyxl')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        # Format the Excel file to make it more readable
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        # Bold the headers
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
        
        # Adjust column widths for better readability
        for col_idx in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 50  # Set a reasonable width
        
        # Save with formatting
        wb.save(output_file)
        print(f"Successfully saved debug information to {output_file}")
    except Exception as e:
        print(f"Warning: Could not apply Excel formatting: {e}")
        print(f"Debug information saved to {output_file} without formatting")


def save_prompt_process(source_text_index,
        source_text,
        target_lang,
        system_prompt,
        prompt,
        model_name,
        review_response):
    # For Debugging: append source text/ relevant specific names/ relevant pair database/ prompt/ response to Excel file
    prompt_file = f'Prompt_Review_PHD_{target_lang}_{model_name}.xlsx'

    import openpyxl
    from openpyxl.styles import Alignment

    try:
        # Check if file exists to determine if we need headers
        file_exists = os.path.isfile(prompt_file)
        
        # Load workbook if it exists, otherwise create a new one
        if file_exists:
            wb = openpyxl.load_workbook(prompt_file)
            ws = wb.active
            # Get next empty row
            next_row = ws.max_row + 1
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            # Add headers if creating a new file
            headers = ["Source Index", "Source Text", "System Prompt", "Prompt", "Response"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            next_row = 2  # Start data from row 2
        
        # Add data to the next row
        ws.cell(row=next_row, column=1, value=str(source_text_index))
        ws.cell(row=next_row, column=2, value=str(source_text))
        ws.cell(row=next_row, column=3, value=str(system_prompt))
        ws.cell(row=next_row, column=4, value=str(prompt))
        ws.cell(row=next_row, column=5, value=str(review_response))

        # Apply word wrap for better readability in Excel
        for col in range(1, 5):
            cell = ws.cell(row=next_row, column=col)
            cell.alignment = Alignment(wrap_text=True)
        
        # Save the workbook
        wb.save(prompt_file)
        
    except Exception as e:
        print(f"Warning: Could not save debug info to Excel: {e}")

def parse_json_column(value):
    """
    Parse a JSON string from Excel back to a Python dictionary/list.
    Ensures proper handling of East Asian characters.
    """
    if isinstance(value, str) and value.strip().startswith('{') and value.strip().endswith('}'):
        try:
            # Use ensure_ascii=False to preserve East Asian characters
            return json.loads(value)
        except json.JSONDecodeError:
            return value
    return value


def make_model_object(review_prompt_obj, model_list, software_type, source_type, source_lang, target_lang, image_path):
    chat_obj_list = []
    # Create LLM chat instance
    
    for model_name in model_list:
        if 'gemini' in model_name:
            # Create Gemini API chat instance
            chat_obj_list.append(
                [
                    GeminiAPIChat(
                        model_name=model_name,
                        system_prompt=review_prompt_obj.sys_prompt,
                        image_path=image_path
                    ),
                ]
            )
                            
        else:
            # Create LLM chat instance
            chat_obj_list.append(
                [
                    OpenaiAPIChat(
                        model_name=model_name,
                        system_prompt= review_prompt_obj.sys_prompt,
                        image_path=image_path
                    ),
                ]
        )

    # print(f'======================Used System Prompt======================')
    # print(chat_obj_list[0].sys_prompt)
    # print(f'======================Used System Prompt======================')
    return chat_obj_list

def get_refer_data(translate_refer, source_text, database_path):
    if translate_refer:
        return translate_refer
    elif not database_path:
        return []
    else:
        relevant_pair_database = search_similar_pair_main(
                translate_dict={"0": source_text}, 
                database_path=database_path, 
                grammar_top_n=5, 
                term_top_n=5
            )
        print(f"Relevant specific names for translation: {relevant_pair_database}")
        return relevant_pair_database

def get_text_group(source_file_path, target_file_path):
    is_xlsx_file = source_file_path.lower().endswith('.xlsx')
    #  Extract data from Excel file
    if is_xlsx_file:
        source_groups = extract_text_from_excel(source_file_path, is_source_file=True)
        target_groups = extract_text_from_excel(target_file_path, is_source_file=False)
        # Find common keys (row indices) that exist in both files
        common_keys = set(source_groups.keys()) & set(target_groups.keys())
        print(f"Found {len(common_keys)} common rows to compare")

    # Extract text data from HTML/ XML files
    else:
        # Determine file type for source and target files
        is_xml_source = source_file_path.lower().endswith('.xml')
        is_xml_target = target_file_path.lower().endswith('.xml')
        source_parser = 'xml' if is_xml_source else 'html.parser'
        target_parser = 'xml' if is_xml_target else 'html.parser'
        print(f"Source file type: {'XML' if is_xml_source else 'HTML'}, using {source_parser} parser")
        print(f"Target file type: {'XML' if is_xml_target else 'HTML'}, using {target_parser} parser")
    
        # Read files with encoding detection
        try:
            # Use the encoding detection function to open files
            source_encoding, source_html = detect_file_encoding(source_file_path)
            target_encoding, html1 = detect_file_encoding(target_file_path)
            
            print(f"Source file encoding: {source_encoding}")
            print(f"Target file encoding: {target_encoding}")
            
        except Exception as e:
            print(f"Error reading HTML/XML files: {e}")
            return
            
        # Parse HTML/XML with BeautifulSoup using appropriate parsers
        bs_source = BeautifulSoup(source_html, source_parser)
        bs1 = BeautifulSoup(html1, target_parser)
        
        # Extract text groups from files
        source_groups = get_text_group_inline(bs_source)
        target_groups = get_text_group_inline(bs1)
    
        # Find groups that exist in both files
        common_keys = set(source_groups.keys()) & set(target_groups.keys())
        print(f"Found {len(common_keys)} common text segments to compare")

    return source_groups, target_groups, common_keys, is_xlsx_file

def truncate_text_for_token_limit(text, max_length=1000):
    """
    Truncates text to fit within token limits.
    For very long texts, tries to preserve the beginning and end which often contain critical context.
    
    Args:
        text (str): The text to truncate
        max_length (int): Maximum allowed length
        
    Returns:
        str: Truncated text
    """
    if len(text) <= max_length:
        return text
    
    # Keep first and last parts, omit middle
    half_length = max_length // 2
    return text[:half_length] + "\n...[Content truncated for token limit]...\n" + text[-half_length:]

def compress_prompt_for_token_limit(prompt, level=1):
    """
    Applies different levels of compression to reduce token usage in prompts.
    
    Args:
        prompt (str): The original prompt
        level (int): Compression level (1-3, with 3 being most aggressive)
        
    Returns:
        str: Compressed prompt
    """
    if level == 1:
        # Level 1: Basic compression - remove unnecessary whitespace and shorten instructions
        compressed = re.sub(r'\s+', ' ', prompt)  # Replace multiple whitespaces with single space
        compressed = compressed.replace("Please provide a detailed analysis", "Analyze")
        compressed = compressed.replace("Please review the following translation", "Review this translation")
        return compressed
    
    elif level == 2:
        # Level 2: Medium compression - truncate examples and lengthy instructions
        compressed = re.sub(r'\s+', ' ', prompt)
        compressed = re.sub(r'For example:.*?(?=\n|$)', 'Be concise.', compressed)
        compressed = re.sub(r'Here are some guidelines.*?(?=\n|$)', '', compressed)
        return compressed
    
    elif level == 3:
        # Level 3: Maximum compression - reduce to core instruction only
        # Extract the essential parts - source text and translation only, with minimal instruction
        source_match = re.search(r'Source Text:.*?(?=\n|$)', prompt)
        source_text = source_match.group(0) if source_match else ""
        
        translation_match = re.search(r'Translation:.*?(?=\n|$)', prompt)
        translation_text = translation_match.group(0) if translation_match else ""
        
        return f"Review concisely: {source_text} {translation_text} List only critical errors."
    
    return prompt  # Default case


async def review_n_improve_process(source_lang,
                                    target_lang,
                                    software_type,
                                    source_type,
                                    source_text, 
                                    translated_text, 
                                    relevant_specific_names,
                                    relevant_region_table,
                                    relevant_refer_text_table,
                                    relevant_pair_database,
                                    image_path,
                                    model_list=None, 
                                    review_chat_obj_list=None, 
                                    improve_chat=None,
                                    temperature=0.3, 
                                    seed=None,
                                    review_path=None,
                                    max_retry_times=1,
                                    need_native_review=False):
    
    print(f'review path is {review_path}')
    
    if need_native_review:
        review_result_dict = {"source_text": source_text, "original_translated_text": translated_text}
        review_result_dict["review_pass_flag"] = False
        review_result_dict["final_translated_text"] = 'Marked as native review, no further review needed.'

    
    else:    
        # 如果沒有提供聊天對象，則創建新的
        review_prompt_obj = ReviewPrompts(source_lang, target_lang, software_type, source_type)
        if not review_chat_obj_list:
            review_chat_obj_list = make_model_object(review_prompt_obj, model_list, software_type, source_type, source_lang, target_lang, image_path)
        
        # if not improve_chat:
        #     improve_chat = OpenaiAPIChat(
        #                 model_name='gpt-4o',
        #                 system_prompt=improve_sys_prompt(source_lang, target_lang, software_type, source_type),
        #                 image_path=image_path
        #             )
        print(f'review_chat_obj_list: {review_chat_obj_list}')
        # print(f'improve_chat_obj_list: {improve_chat}')
        print(f'Get translated text: {translated_text}')

        try:
            review_result_dict = {"source_text": source_text, "original_translated_text": translated_text}
            review_prompt_obj.source_text = source_text
            review_prompt_obj.translation = translated_text
            review_prompt_obj.specific_names = relevant_specific_names
            review_prompt_obj.region_table = relevant_region_table
            review_prompt_obj.refer_text_table = relevant_refer_text_table
            review_prompt_obj.translate_refer = relevant_pair_database
            reviewed_dict = {}
            # improved_dict = {0: translated_text}
            process_pass_flag = False

            for retry_time in range(max_retry_times):
                print(f'Current Doing {retry_time+1} times translation...')
                
                for _ in range(len(model_list)):
                    model_name = model_list[_]
                    print(f'========================Used Model: {model_name}========================')
                    check_item_index_dict = {
                            0: 'all'
                        }
  
                    raw_review_response_dict ={}
                    for check_item_index in range(len(review_chat_obj_list[_])):
                        print(f'===========Checking Point: {check_item_index_dict[check_item_index]}===========')

                        review_chat = review_chat_obj_list[_][check_item_index]
                        
                        if 'o3' in model_name:
                            kwargs = {}
                        else:
                            kwargs = {"temperature": temperature}
                        if seed is not None:
                            kwargs["seed"] = seed

                        # First - do the review
                        review_response = ''
                        review_stop_reason = ''
                        if check_item_index == 0:
                            prompt_text = review_prompt_obj.review_prompt()


                        # print('='*40)
                        # print(f'Current review prompt:\n{prompt_text}')
                        # print('='*40)

                        # Handle possible token limitation
                        try:
                            async for chunk, review_stop_reason in review_chat.get_stream_aresponse(prompt_text, **kwargs):
                                review_response += chunk
                                
                            if review_stop_reason == 'length':
                                print("Review response exceeded length limit but received partial content.")
                                raise RuntimeError("Review response too short after hitting length limit.")
                            
                        except RuntimeError as e:
                            print(f"Review process failed: {str(e)}")
                            raise RuntimeError("Translation review failed due to length limit or other issues.")
                        print(f"Review response:\n {review_response}")
                        
                        # Parse the review response
                        review_response_json = as_json_obj(review_response)
                        raw_review_response_dict[check_item_index_dict[check_item_index]] = review_response_json

                        # save_prompt_process(
                        #     source_text_index=0,
                        #     source_text=source_text,
                        #     target_lang=target_lang,
                        #     system_prompt=review_prompt_obj.sys_prompt,
                        #     prompt=prompt_text,
                        #     model_name=model_name,
                        #     review_response=review_response
                        # )

                    print(f"Raw review response dictionary for {retry_time+1} times: {raw_review_response_dict}")
                    
                    # Export debug information
                    debug_file = "debug_review.xlsx"
                    # debug_file = f"Prompt_Review_{target_lang}_{model_name}_PHD.xlsx"
                    try:
                        export_debug_info(
                            source_text=source_text,
                            translated_text=translated_text,
                            relevant_specific_names=relevant_specific_names,
                            relevant_region_table=relevant_region_table,
                            relevant_refer_text_table=relevant_refer_text_table,
                            relevant_pair_database=relevant_pair_database,
                            review_prompt_obj=review_prompt_obj,
                            raw_review_response_dict=raw_review_response_dict,
                            review_result_dict=review_result_dict,
                            model_name=model_name,
                            output_file=debug_file
                        )
                        print(f"Exported debug information to {debug_file}")
                    except Exception as e:
                        print(f"Failed to export debug information: {e}")

                    if retry_time+1 not in reviewed_dict.keys():
                        reviewed_dict[retry_time+1] = {model_name: None}

                    if all (value is None for value in raw_review_response_dict.values()):
                        print("Review response is empty or invalid JSON, attempting to extract useful information.")
                        # Try to salvage some information from non-JSON response
                        process_pass_flag = f'Error in review response with {model_name}'
                        reviewed_dict[retry_time+1][model_name] = review_response
                        continue
                            
                    else:
                        # # Check if the key exists in the dictionary
                        # if key not in reviewed_dict[retry_time+1][model_name]:
                        #     reviewed_dict[retry_time+1][model_name] = {}
                        review_response_dict = {}
                        try:
                            for key, value in raw_review_response_dict.items():
                                if value is None or not isinstance(value, dict):
                                    review_response_dict[key] = None
                                else:
                                    review_response_dict[key] = value['Suggestions']
                            # Store the improvement suggestions
                            reviewed_dict[retry_time+1][model_name] = review_response_dict
                            print(f'reviewed_dict for {retry_time+1} times: {reviewed_dict}')
                        except Exception as e:
                            reviewed_dict[retry_time+1][model_name] = raw_review_response_dict
                            print(f"Error processing review response for {model_name}: {str(e)}")
                            process_pass_flag = f'Error in review response with {model_name}'

                print(f'Current reviewed_dict: {reviewed_dict}')

                if type(process_pass_flag) == str and 'Error in review response' in process_pass_flag:
                    print("Error in review response, skipping re-translation.")
                    break
                
                if (retry_time+1) in reviewed_dict and all(value is None for val in reviewed_dict[retry_time+1].values() for value in val.values()):
                    print(f"All models returned None for review in attempt {retry_time+1}, skipping re-translation.")
                    process_pass_flag = True
                    break

            #     # Second - do the re-translation
            #     improve_response = ''
            #     improve_stop_reason = ''
                

            #     # Combine the review suggestions into a single string
            #     suggestions = []
            #     for model_name, suggestions_list in reviewed_dict[retry_time+1].items():
            #         if isinstance(suggestions_list, str):
            #             suggestions.append(suggestions_list)
            #         elif isinstance(suggestions_list, list):
            #             suggestions.extend(suggestions_list)
            #         elif isinstance(suggestions_list, dict):
            #             suggestions.extend(list(suggestions_list.values()))
            #     # suggestions = [s.strip() for s in suggestions if s.strip()]
            #     print(f"Suggestions for re-translation: {suggestions}")

            #     # Combine translated text (improved_dict) to a list
            #     translated_text_list_str = str(list(improved_dict.values()))
            #     print(f"Translated text list for re-translation: {translated_text_list_str}")

            #     improve_text = improve_prompt(source_lang, target_lang, source_text, translated_text, relevant_specific_names, relevant_pair_database, suggestions=suggestions, translated_text=translated_text_list_str)
                
            #     # print(f'='*40)
            #     # print(f'Current improve_text:\n{improve_text}')
            #     # print(f'='*40)
            #     try:
            #         async for chunk, improve_stop_reason in improve_chat.get_stream_aresponse(improve_text, temperature=0.01):
            #             improve_response += chunk
                    
            #         if improve_stop_reason == 'length':
            #             raise RuntimeError("Improve response too short after hitting length limit.")
                    
            #     except RuntimeError as e:
            #         print(f"Improve process failed: {str(e)}")
                
            #     print(f"Improve raw response:\n {improve_response}")
                
            #     # Parse the re-translation response                
            #     improve_json = as_json_obj(improve_response)
            #     if not improve_json:
            #         print("Improve response is not in expected JSON format, trying to extract translation.")
            #         translated_text = improve_response
            #         improved_dict[retry_time+1] = translated_text
            #         process_pass_flag = 'Error in improve response'
            #         break

            #     else:
            #         translated_text = list(improve_json.values())[-1]
            #         improved_dict[retry_time+1] = translated_text

            #     if retry_time == 2:
            #         process_pass_flag = False

            # if 2 not in reviewed_dict.keys(): reviewed_dict[2] = {}
            # if 3 not in reviewed_dict.keys(): reviewed_dict[3] = {}

            # for model_name in model_list:
            #     if model_name not in reviewed_dict[2].keys():
            #         reviewed_dict[2][model_name] = 'N/A'
            #     if model_name not in reviewed_dict[3].keys():
            #         reviewed_dict[3][model_name] = 'N/A'

            # if 1 not in improved_dict.keys(): improved_dict[1] = 'N/A'
            # if 2 not in improved_dict.keys(): improved_dict[2] = 'N/A'
            # if 3 not in improved_dict.keys(): improved_dict[3] = 'N/A'

            # print(f'Current review result: {reviewed_dict}')
            # print(f'Current improved result: {improved_dict}')            # Create a summary dictionary to collect all suggestions for each model
            summary_by_model = {}
            
            for key, value in reviewed_dict.items():
                for model_name, suggestions in value.items():
                    # Format suggestions dictionary with proper indentation if it's a dictionary
                    if isinstance(suggestions, dict):
                        import json
                        # Store as a formatted JSON string with indentation
                        # ensure_ascii=False preserves East Asian characters
                        review_result_dict[f"{model_name}_review_{key}"] = json.dumps(suggestions, indent=4, ensure_ascii=False)
                        
                        # Add to the model's summary if it's not already there
                        if model_name not in summary_by_model:
                            summary_by_model[model_name] = []
                        
                        # Add only the suggestion content to the summary
                        for category, suggestion in suggestions.items():
                            if suggestion:  # Only add non-empty suggestions
                                summary_by_model[model_name].append(f"{category}: {suggestion}")
                    else:
                        review_result_dict[f"{model_name}_review_{key}"] = suggestions
                
                # review_result_dict[f"improved_{key}"] = improved_dict[key]
            
            # Add the summary columns to the review result dictionary
            for model_name, summary_items in summary_by_model.items():
                if summary_items:
                    # Join all suggestions with line breaks                    
                    summary_text = "\n".join(summary_items)
                    review_result_dict[f"summary_{model_name}"] = summary_text
                else:
                    review_result_dict[f"summary_{model_name}"] = "No suggestions"
            
            review_result_dict["review_pass_flag"] = process_pass_flag
            review_result_dict["final_translated_text"] = translated_text

        except Exception as e:
            error_message = str(e)
            print(f"{error_message}")
            return error_message
        
    try:
        print('='*40)
        print(f'review result dict: {review_result_dict}')
        print('='*40)
        # Append review_result_dict to the "review_results.xlsx" file
        if review_path:
            # Load existing results if the file exists, otherwise create a new DataFrame
            if os.path.exists(review_path):
                try:
                    # Ensure we use openpyxl engine which has better East Asian character support
                    existing_df = pd.read_excel(review_path, engine='openpyxl')
                    # Parse any JSON strings in the DataFrame
                    for col in existing_df.columns:
                        if "_review_" in col:
                            existing_df[col] = existing_df[col].apply(parse_json_column)
                except Exception as e:
                    print(f"Error reading {review_path}: {e}")
                    existing_df = pd.DataFrame()
            else:
                existing_df = pd.DataFrame()

            # Convert the review result into a DataFrame row
            # Ensure proper handling of East Asian characters
            for key, value in review_result_dict.items():
                if isinstance(value, str):
                    # Ensure strings are properly encoded for Excel
                    review_result_dict[key] = value
            
            new_df = pd.DataFrame([review_result_dict])

            # Append the new data
            final_df = pd.concat([existing_df, new_df], ignore_index=True)

            # Save the updated results back to the Excel file
            final_df.to_excel(review_path, index=False, engine='openpyxl')
            
            # Reload with openpyxl to ensure proper encoding of East Asian characters
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment
                
                # Load the workbook
                wb = openpyxl.load_workbook(review_path)
                ws = wb.active
                
                # Apply formatting to make the file more readable
                for col_idx, column in enumerate(final_df.columns):
                    # Bold header
                    cell = ws.cell(row=1, column=col_idx+1)
                    cell.font = Font(bold=True)
                
                # Save with formatting
                wb.save(review_path)
                print(f"Successfully saved and formatted review results to {review_path}")
            except Exception as e:
                print(f"Warning: Could not apply Excel formatting: {e}")
        return translated_text
    except Exception as e:
        error_message = str(e)
        print(f"Error saving review results: {error_message}")
        return error_message



# 修改後的 compare_result 函數，修復了事件循環問題
async def process_segments(
        source_groups,
        target_groups,
        common_keys,
        is_xlsx_file,
        model_list,
        software_type,
        source_type,
        source_lang,
        target_lang,
        specific_names,
        region_table,
        refer_text_table,
        translate_refer,
        database_path,
        temperature,
        seed,
        review_report_path
):
    review_results = []
    
    # # 建立模型物件，在循環外部只建立一次
    # review_chat_obj_list = make_model_object(model_list, software_type, source_type, source_lang, target_lang, image_path=None)

    for i, key in enumerate(sorted(common_keys, key=lambda x: int(x))):
        print(f"Comparing segment {i+1}/{len(common_keys)}")
        
        # Get the corresponding group from each file
        if is_xlsx_file:
            source_text = source_groups.get(key, "Source text not available")
            translated_text = target_groups.get(key)
        else:
            source_group = source_groups.get(key)
            group1 = target_groups[key]
            
            # Get text content
            source_text = str(source_group) if source_group else "Source text not available"
            translated_text = str(group1)
            
        print(f'Current is doing source text {source_text}')
        
        # Identify specific named entities in the current segment
        relevant_specific_names = get_relevant_specific_names(specific_names, source_text)

        # Get the region table
        relevant_region_table = get_relevant_region_table(region_table, source_text)
        print(f"Relevant region table for translation: {relevant_region_table}")

        # Get the refer text table
        relevant_refer_text_table = get_relevant_refer_text_table(refer_text_table, source_text)
        print(f"Relevant refer text table for translation: {relevant_refer_text_table}")


        # Create the prompt
        relevant_pair_database = get_refer_data(translate_refer, source_text, database_path)
        
        # 使用 await 而不是 asyncio.run() 來呼叫 review_n_improve_process
        result = await review_n_improve_process(
            source_lang=source_lang,
            target_lang=target_lang,
            software_type=software_type,
            source_type=source_type,
            source_text=source_text,
            translated_text=translated_text,
            relevant_specific_names=relevant_specific_names,
            relevant_region_table=relevant_region_table,
            relevant_refer_text_table=relevant_refer_text_table,
            relevant_pair_database=relevant_pair_database,
            image_path=None,
            model_list=model_list,
            # review_chat_obj_list=review_chat_obj_list,
            # improve_chat=improve_chat,
            temperature=temperature,
            seed=seed,
            review_path=review_report_path
        )
        
        review_results.append(result)
    
    return review_results

def compare_result(
        source_file_path: str,
        target_file_path: str, 
        output_path_list: list,
        model_list: list,
        software_type: str,
        specific_names: dict = None,
        region_table: dict = None,
        refer_text_table: dict = None,
        temperature: float = 0.3,
        seed: int = None,
        source_lang: str = conf.SOURCE_LANGUAGE,
        target_lang: str = conf.TARGET_LANGUAGE,
        source_type: str = conf.SOURCE_TYPE,
        translate_refer: list = None,
        database_path: str = None,
        review_report_path: str = None
) -> None:
    """
    Compare two translated HTML files and generate a review report in HTML format.
    
    :param source_file_path: Path to the original source file (source language)
    :param target_file_path: Path to the first HTML file (target language)
    :param output_path_list: Path to save the review report HTML
    :param mode_list: List of modes to use for review (e.g., 'UI', 'Help', etc.)
    :param specific_names: Dictionary of specific terms to translate in a specific way
    :param temperature: Temperature parameter for controlling randomness (0.0-2.0, lower is more deterministic)
    :param seed: Optional seed value for reproducible results
    :param source_lang: Source language (e.g., 'English')
    :param target_lang: Target language (e.g., 'Traditional Chinese')
    :param source_type: Type of source file (e.g., 'UI', 'Help', etc.)
    :param software_type: Type of software being translated (e.g., 'video editing software', 'image editing software', etc.)
    :param translate_refer: List of references for translation
    :param database_path: Path to the database for translation references
    :return: None
    """
    print(f"Starting review using source: {source_file_path}")
    print(f"Comparing source file with target file: {source_file_path} and {target_file_path}")
    print(f"Using source language: {source_lang}")
    print(f"Using target language: {target_lang}")
    print(f"Using software type: {software_type}")
    print(f"Using source type: {source_type}")
    
    source_groups, target_groups, common_keys, is_xlsx_file = get_text_group(source_file_path, target_file_path)
    print(f'Source groups: {len(source_groups)} segments found')
    
    # Initialize review results
    review_results = []
    for model in model_list:
        review_results.append([])
    
    # 正確設置和管理事件循環
    print("Comparing segments...")
    try:
        # 使用一個新的事件循環
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        # 運行非同步處理函數
        results = loop.run_until_complete(
            process_segments(
                source_groups,
                target_groups,
                common_keys,
                is_xlsx_file,
                model_list,
                software_type,
                source_type,
                source_lang,
                target_lang,
                specific_names,
                region_table,
                refer_text_table,
                translate_refer,
                database_path,
                temperature,
                seed,
                review_report_path
            )
        )
        
        # 最後關閉事件循環
        loop.close()
        
        # 處理結果
        review_results.extend(results)
        
    except Exception as e:
        print(f"Error during review: {e}")
    
    # 可以在這裡處理和保存比較結果
    print("review processing completed")


def main(
        input_file_path="default", 
         output_file_path="default", 
         compare_file_path="default", 
         specific_names_xlsx_path="default",
         region_table_path="default",
         refer_text_table_path="default",
         software_type="default", 
         source_lang="default", 
         target_lang="default",
         source_type="default",
         translate_refer="default",
         database_path="default",
         model_list="default",
         review_report_path="default"):
    
    """Command-line entry point for review functionality"""
    if input_file_path=="default":
        input_file_path = conf.INPUT_FILE_PATH
    if output_file_path=="default":
        output_file_path = conf.OUTPUT_FILE_PATH
    if compare_file_path=="default":
        compare_file_path = conf.COMPARE_FILE_PATH
    if specific_names_xlsx_path=="default":
        specific_names_xlsx_path = conf.SPECIFIC_NAMES_XLSX
    if region_table_path=="default":
        region_table_path = conf.REGION_TABLE_PATH
    if refer_text_table_path=="default":
        refer_text_table_path = conf.REFER_TEXT_TABLE_PATH
    if software_type=="default":
        software_type = conf.SOFTWARE_TYPE
    if source_lang=="default":
        source_lang = conf.SOURCE_LANGUAGE
    if target_lang=="default":
        target_lang = conf.TARGET_LANGUAGE
    if source_type=="default":
        source_type = conf.SOURCE_TYPE
    if database_path=="default":
        database_path = conf.DATABASE_PATH
    if model_list=="default":
        model_list = conf.COMPARISON_MODEL
    print("Running in review mode...")
    print(f"Comparing Source file: {input_file_path}")
    print(f"Comparing Translated file: {output_file_path}")
    print(f"Output review base path: {compare_file_path}")
    print(f"Using software type: {software_type}")
    print(f"Using source language: {source_lang}")
    print(f"Using target language: {target_lang}")
    print(f"Using source type: {source_type}")
    print(f"Using database path: {database_path}")
    
    # Load specific names if configured
    specific_names = {}
    if specific_names_xlsx_path:
        try:
            specific_names = load_specific_names(specific_names_xlsx_path, source_lang, target_lang)
            print(f"Loaded {len(specific_names)} specific name translations for review")
        except Exception as e:
            print(f"Warning: Could not load specific names: {e}")
    
    region_table = {}
    if region_table_path:
        try:
            region_table = load_region_table(region_table_path, source_lang, target_lang)
            print(f"Loaded {len(region_table)} region translations for review")
        except Exception as e:
            print(f"Warning: Could not load region table: {e}")
    
    refer_text_table = {}
    if refer_text_table_path:
        try:
            refer_text_table = load_refer_text_table(refer_text_table_path, source_lang, target_lang)
            print(f"Loaded {len(refer_text_table)} reference translations for review")
        except Exception as e:
            print(f"Warning: Could not load refer text table: {e}")
    
    # Get temperature and seed from config if available
    temperature = getattr(conf, 'TEMPERATURE', 0.3)
    seed = getattr(conf, 'SEED', None)
    
    if temperature != 0.3:
        print(f"Using temperature: {temperature}")
    if seed is not None:
        print(f"Using seed: {seed}")
      # Check file extensions to determine file type

    # Create model-specific output file path by appending model name to the filename
    file_base, file_ext = os.path.splitext(compare_file_path)
    model_output_path_list = []
    for model_name in model_list:
        model_output_path_list.append(f"{file_base}_{model_name.replace('-', '_')}{file_ext}")
    
    print(f"Output will be saved to: {model_output_path_list}")
    
    # Run the appropriate review based on file types
    compare_result(
        input_file_path,
        output_file_path,
        model_output_path_list,
        model_list,
        software_type,
        specific_names,
        region_table,
        refer_text_table,
        temperature=temperature,                
        seed=seed,
        source_lang=source_lang,
        target_lang=target_lang,
        source_type=source_type,
        translate_refer=translate_refer,
        database_path=database_path,
        review_report_path=review_report_path
    )

    print(f"review completed")


if __name__ == '__main__':
    main()
